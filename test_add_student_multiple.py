import requests
import openpyxl
import json

def test_add_multiple_students():
    URL = "https://thetestingworldapi.com/api/studentsDetails"
    f = open("request_json.json", "r")
    json_request = json.loads(f.read())
    workbook = openpyxl.load_workbook("students.xlsx")
    sheet = workbook["Munka1"]
    rows = sheet.max_row

    for i in range(2, rows+1):
        cell_first_name = sheet.cell(i, 1)
        cell_middle_name = sheet.cell(i, 2)
        cell_last_name = sheet.cell(i, 3)
        cell_dob = sheet.cell(i, 4)
        json_request["first_name"] = cell_first_name.value
        json_request["middle_name"] = cell_middle_name.value
        json_request["last_name"] = cell_last_name.value
        json_request["date_of_birth"] = cell_dob.value
        response = requests.post(URL, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201