# This file will dynamically update the json request based on the excel.

import json
import openpyxl
import requests

class ReaderWriter:

    def __init__(self, file_name_path, sheetname):
        global workbook
        global sheet
        workbook = openpyxl.load_workbook(file_name_path)
        sheet = workbook[sheetname]
    
    def fetch_key_names(self):
        col = sheet.max_column
        key_list = []
        for i in range(1, col+1):
            cell = sheet.cell(1, i)
            key_list.append(cell.value)
        return key_list
    
    def update_request(self, row_number, json_request, key_list):
        col = sheet.max_column
        for i in range(1, col+1):
            cell = sheet.cell(row_number, i)
            json_request[key_list[i-1]] = cell.value
        return json_request

def test_add_multiple_students():
    URL = "https://thetestingworldapi.com/api/studentsDetails"
    f = open("request_json.json", "r")
    json_request = json.loads(f.read())

    reader_writer = ReaderWriter("students.xlsx", "Munka1")
    key_list = reader_writer.fetch_key_names()
    rows = sheet.max_row
    print()
    for i in range(2, rows+1):
        updated_json_req = reader_writer.update_request(i, json_request, key_list)
        response = requests.post(URL, updated_json_req)
        print(response.text)
    
   